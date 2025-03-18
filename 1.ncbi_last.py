from Bio import Entrez, SeqIO
from io import StringIO
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# NCBI Entrez Email 등록
Entrez.email = "wijae0523@orcid"

# 파일 경로 설정
file_path = 'clinvar_mutation_file.xlsx'
output_file_path = 'last_data_corrected_highlighted.xlsx'

# 엑셀 파일 읽기
df = pd.read_excel(file_path)

# 열 이름 정규화 (대소문자 및 공백 제거)
df.columns = df.columns.str.strip()

# 'Canonical SPDI' 열 존재 여부 확인
if 'Canonical SPDI' not in df.columns:
    raise ValueError("The required column 'Canonical SPDI' is missing from the input file.")

# 결과 저장용 리스트 생성
results = []

total_rows = len(df)
for index, row in df.iterrows():
    try:
        print(f"Processing row {index + 1} of {total_rows}...")

        # 기존 열 데이터 추출
        name = row.get('Name', 'Unknown')  # Name 열이 없는 경우 'Unknown'으로 처리
        chromosome = row.get('GRCh38Chromosome', 'Unknown')

        # Canonical SPDI 값에서 RefSeq ID와 위치 추출
        canonical_spdi = row['Canonical SPDI']

        # 결측값 또는 비문자열 처리
        if not isinstance(canonical_spdi, str) or pd.isna(canonical_spdi):
            print(f"Skipping row {index + 1} due to invalid Canonical SPDI value.")
            continue

        refseq_id, location, original_base, new_base = canonical_spdi.split(':')
        location = int(location) + 1  # 0-기반 좌표계를 1-기반으로 변환

        # 위치 정보에서 시작과 끝 설정 (50bp window 예시)
        start = max(location - 25, 1)
        stop = location + 25

        # NCBI에서 서열 가져오기
        handle = Entrez.efetch(
            db="nucleotide",
            id=refseq_id,
            seq_start=start,
            seq_stop=stop,
            rettype="fasta",
            retmode="text"
        )
        response = handle.read()
        handle.close()

        # FASTA 형식 확인 및 서열 읽기
        if response.startswith(">"):
            record = SeqIO.read(StringIO(response), "fasta")
            sequence = str(record.seq)

            # 결과 저장
            results.append({
                "Name": name,
                "Canonical SPDI": canonical_spdi,
                "GRCh38Chromosome": chromosome,
                "RefSeqID": refseq_id,
                "Position": location,  # 수정된 위치 값 (1-기반)
                "Start": start,
                "Stop": stop,
                "Sequence": sequence,
                "OriginalBase": original_base,
                "NewBase": new_base
            })
            print(f"Row {index + 1} processed successfully.")
        else:
            print(f"Invalid FASTA format for row {index + 1}.")
    except Exception as e:
        print(f"Error processing row {index + 1}: {e}")

# 결과를 데이터프레임으로 변환
df_results = pd.DataFrame(results)

# 엑셀 파일로 저장
df_results.to_excel(output_file_path, index=False, engine='openpyxl')

# 강조된 서열 추가
wb = load_workbook(output_file_path)
ws = wb.active

# 새로운 시트 추가
highlight_sheet = wb.create_sheet(title="Highlighted Sequences")
highlight_sheet.append(["Name", "GRCh38Chromosome", "Position", "Canonical SPDI", "Original Sequence", "Highlighted Sequence"])  # 헤더 추가

for index, row in df_results.iterrows():
    seq = row['Sequence']
    highlight_index = 25  # 중앙 위치 기준 강조
    name = row['Name']
    canonical_spdi = row['Canonical SPDI']
    chromosome = row['GRCh38Chromosome']
    position = row['Position']

    # 강조 표시된 텍스트 생성
    formatted_seq = ""
    for i, char in enumerate(seq):
        if i == highlight_index:
            formatted_seq += f"[{char}]"  # 강조 표시 부분
        else:
            formatted_seq += char

    # 기존 시트에 강조 표시 업데이트
    ws[f"H{index + 2}"] = formatted_seq  # H 열에 강조된 시퀀스를 추가
    ws[f"H{index + 2}"].font = Font(color="FF0000")

    # 새로운 시트에 데이터 추가
    highlight_sheet.append([name, chromosome, position, canonical_spdi, seq, formatted_seq])

# 저장
wb.save(output_file_path)
print(f"Highlighted sequences saved to {output_file_path}")
